"""Exportación de inventario a Excel (.xlsx)."""

import io

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def nombre_usuario(usuario) -> str:
    if not usuario:
        return ""
    nombre = f"{usuario.first_name or ''} {usuario.last_name or ''}".strip()
    return nombre or (usuario.email or "")


def exportar_activos_excel(activos, filename: str, filtros_aplicados=None) -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Activos"

    header_fill = PatternFill("solid", fgColor="32407B")
    header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    title_font = Font(color="32407B", bold=True, name="Calibri", size=14)
    thin = Border(
        left=Side(style="thin", color="D0D5DD"),
        right=Side(style="thin", color="D0D5DD"),
        top=Side(style="thin", color="D0D5DD"),
        bottom=Side(style="thin", color="D0D5DD"),
    )
    zebra = PatternFill("solid", fgColor="F8F9FC")

    headers = [
        "Código inventario",
        "Categoría",
        "Subcategoría",
        "Marca",
        "Modelo",
        "Nº serial",
        "Estado",
        "Ubicación",
        "Usuario asignado",
        "Observaciones",
        "Fecha creación",
        "Fecha actualización",
    ]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"] = "Inventario de Activos — Consorcio PALDACA"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    generado = timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M")
    meta_parts = [f"Generado: {generado}"]
    if filtros_aplicados:
        meta_parts.append(
            "Filtros: " + ", ".join(f"{k}: {v}" for k, v in filtros_aplicados.items())
        )
    meta_parts.append(f"Total: {len(activos)}")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws["A2"] = " | ".join(meta_parts)
    ws["A2"].font = Font(color="667085", name="Calibri", size=9)

    header_row = 4
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    ws.row_dimensions[header_row].height = 28
    ws.freeze_panes = f"A{header_row + 1}"

    for row_idx, activo in enumerate(activos, start=header_row + 1):
        values = [
            activo.codigo_inventario or "",
            activo.subcategoria.categoria.nombre if activo.subcategoria_id else "",
            activo.subcategoria.nombre if activo.subcategoria_id else "",
            activo.marca or "",
            activo.modelo or "",
            activo.numero_serial or "",
            activo.get_estado_display(),
            activo.ubicacion.nombre if activo.ubicacion_id else "",
            nombre_usuario(activo.usuario_asignado),
            activo.observaciones or "",
            activo.fecha_creacion.strftime("%d/%m/%Y %H:%M") if activo.fecha_creacion else "",
            (
                activo.fecha_actualizacion.strftime("%d/%m/%Y %H:%M")
                if activo.fecha_actualizacion
                else ""
            ),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if (row_idx - header_row) % 2 == 0:
                cell.fill = zebra

    last_data_row = header_row + max(len(activos), 1)
    ws.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(len(headers))}{last_data_row}"
    )
    widths = [18, 16, 16, 14, 16, 16, 16, 16, 22, 28, 16, 16]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
